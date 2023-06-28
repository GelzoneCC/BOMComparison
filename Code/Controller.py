# -*- coding: utf-8 -*-
"""
Created on Tue May 23 15:20:23 2023

@author: GelzoneCC
"""

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog, QMainWindow
from AgileCompareRD_UI3 import Ui_AgileRDMainWindow
import time, win32api, shutil
from decimal import Decimal, ROUND_HALF_UP

startTime = time.time()

###----------Application UI initialization, flow control.----------###
class MainWindowController(QMainWindow, Ui_AgileRDMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        #self.setFixedSize(560, 330)
        self.setupControl()
        self.progressBar.setStyleSheet( #Set CSS-like format for progress bar.
            '''
            QProgressBar {
                border: 2px solid #000;
                background:#aaa;
                color:#fff;
                border-radius: 8px;
                text-align:center;
                }
            QProgressBar::chunk {
                background: #333;
                width:1px;
                }
            '''
            )

    def setupControl(self):
        self.rdBrowseBtn.clicked.connect(self.openRDBOM)
        self.agileBrowseBtn.clicked.connect(self.openAgileBOM)
        self.changeListBtn.clicked.connect(self.changeList)
    
    def openRDBOM(self):
        rdBOMName, fileType = QFileDialog.getOpenFileName(self, "Open RD BOM", "{}".format(os.path.expanduser('~/Documents'))) #Browse file window title: Open RD BOM.
        self.rdFileNameText.setPlainText(rdBOMName)
    
    def openAgileBOM(self):
        agileBOMName, fileType = QFileDialog.getOpenFileName(self, "Open Agile BOM", "{}".format(os.path.expanduser('~/Documents')))
        self.agileFileNameText.setPlainText(agileBOMName)
        
    def msgBox(self, title, content, icon):
        self.msg = QtWidgets.QMessageBox(self)
        self.msg.setWindowTitle(title)
        self.msg.setText(content)
        self.msg.setIcon(icon)
        self.msg.addButton(QtWidgets.QMessageBox.Ok)
        #self.msg.setDefaultButton(QtWidgets.QMessageBox.Ok)
        self.msg.exec()
    
    #----------Generate change list.----------
    def changeList(self):
        if self.rdFileNameText.toPlainText() == '' or self.agileFileNameText.toPlainText() == '' or self.configInputText.toPlainText() == '': #Check whether files are empty.
            self.msgBox('Warning!', 'Please upload files or input config.', 2) #Warning.
        elif not self.rdFileNameText.toPlainText().lower().endswith('.xlsm') or not self.agileFileNameText.toPlainText().lower().endswith('.xls'): #Check whether input the correct files.
            self.msgBox('Warning!', 'Please confirm whether you have uploaded the correct files.', 2)
        elif self.configInputText.toPlainText() not in getFullConfig(self.rdFileNameText.toPlainText()): #Check whether input correct config.
            self.msgBox('Warning', 'The config does not show in the RD BOM, please check it.', 2)
        else:
            self.progressBar.setValue(0)
            if os.path.exists(backupPath) == False:
                os.mkdir(backupPath)
            if os.path.exists(backupPath + '/Input') == False:
                os.mkdir(os.path.join(backupPath, 'Input'))
            if os.path.exists(backupPath + '/Output') == False:
                os.mkdir(os.path.join(backupPath, 'Output'))
            shutil.copy(self.rdFileNameText.toPlainText(), backupPath + '/Input') #Copy input file to file server.
            shutil.copy(self.agileFileNameText.toPlainText(), backupPath + '/Input')
            initChangeList()
            toChangeList(agileBOM(self.agileFileNameText.toPlainText()), rdSMBOM(self.rdFileNameText.toPlainText(), self.configInputText.toPlainText()), self.configInputText.toPlainText())
            self.msgBox('Completed!', 'Change List has been generated in the current path.\n{}'.format(outputPath + '/' + currTime + '_Agile update.xlsx'), 1) #Information.
            #self.rdFileNameText.setPlainText('')
            #self.agileFileNameText.setPlainText('')
            self.configInputText.setPlainText('')
            self.progressBar.reset()
            win32api.ShellExecute(0, 'open', os.path.dirname(self.rdFileNameText.toPlainText()) + '/Output' + '/' + currTime + '_Agile update.xlsx', '', '', 0) #Open file after execution.
            shutil.copy(outputPath + '/' + currTime + '_Agile update.xlsx', backupPath + '/Output')
            self.exeTime = time.strftime("%M:%S", time.gmtime(time.time() - startTime))
            logList.append('Execution time: {}.'.format(self.exeTime))
            logList.append('Completed. Please refer to the following path for the agile update result.\n{}'.format(os.path.dirname(self.rdFileNameText.toPlainText()) + '/' + currTime + '_Agile update.xlsx'))
            logFile()

###----------Move data process code here to connect UI flow control.----------###
import openpyxl, xlsxwriter, xlrd, os, pylightxl, copy
import pandas as pd
import numpy as np
from anytree import Node, RenderTree, PreOrderIter


###----------agile BOM----------###
def agileBOM(filePath):
    #Read agile system BOM as a dataframe.
    agileBOM_df = pd.read_excel(filePath, header = 0, index_col = None, engine = 'xlrd')
    agileBOM_df.loc[agileBOM_df['Level'] == 0, 'BOM.Qty'] = 1 #Set BOM.Qty = 1 for all level 0 in agile BOM.
    agileBOM_df['BOM.Qty'].convert_dtypes()
    agileBOM_df['BOM.Qty'] = agileBOM_df['BOM.Qty'].astype('Int64') #Convert to pandas integer including NaN value.
    #Remove rows with blank based on Number.
    agileBOM_df['Number'].replace('', np.nan, inplace=True)
    agileBOM_df.dropna(subset = ['Number'], inplace = True)
    #Plus 1 in Level.
    agileBOM_df['Level'].convert_dtypes()
    agileBOM_df['Level'] += 1
    agileBOM_df['Level'] = agileBOM_df['Level'].astype(int)
    #Remove rows with '864KE' in Number.
    agileBOM_df.drop(agileBOM_df[agileBOM_df.Number == 'xxxxx'].index, inplace = True)
    #Keep key column.
    agileBOMKeyCol_df = agileBOM_df[['Level', 'Number', 'BOM.Qty', '*Description']] #Remove writer and add key column df.
    agileBOMKeyCol_df['*Description'] = agileBOMKeyCol_df['*Description'].str.rstrip() #Remove blank at the end of the description.
    agileBOMKeyCol_df = agileBOMKeyCol_df.reset_index(drop = True)
    
    return agileBOMKeyCol_df

###----------RD SMBOM----------###
def rdSMBOM(filePath, config):
    #----------Use pylightxl to read RD SMBOM as a dataframe and keep key column.----------
    wb = pylightxl.readxl(fn = filePath, ws = 'BOM')
    data = []
    #Ignore the button in cells by just reading the data we need.
    for row in wb.ws('BOM').rows:
        rowData = []
        for cell in row:
            if isinstance(cell, str): #Add to rowData if the cell contains text.
                rowData.append(cell)
            elif isinstance(cell, int):
                rowData.append(cell)
            elif isinstance(cell, float):
                rowData.append(cell)
        data.append(rowData)
    rdSMBOM_df = pd.DataFrame(data[1:], columns = data[0])
    rdSMBOM_df['Unique Identifier'] = rdSMBOM_df['Unique Identifier'].str.rstrip()
    rdSMBOM_df['DPN'].replace('', np.nan, inplace = True) #Convert empty string to NaN.
    rdSMBOM_df['DPN'].replace(' ', np.nan, inplace = True) #most null value is a space in SMBOM...
    
    #Find index of LVL == 0's row and the next index of LVL == 0's row to grab the data between them.
    motherLevelIndex = rdSMBOM_df[(rdSMBOM_df['LVL'] == 0) & (rdSMBOM_df['Unique Identifier'] == config)].index[0]
    if rdSMBOM_df.loc[motherLevelIndex + 1:, 'LVL'].eq(0).any(): #Check if there's next LVL == 0's row.
        nextLevelIndex = rdSMBOM_df[rdSMBOM_df.index > motherLevelIndex]['LVL'].idxmin() #Get index of next LVL == 0's row.
        rowsBetween = rdSMBOM_df.loc[motherLevelIndex + 1:nextLevelIndex - 1] #Get data between 2 rows.
    else:
        rowsBetween = rdSMBOM_df.loc[motherLevelIndex + 1:]
    rowsBetween.dropna(subset = ['DPN'], inplace = True)
    #Keep key column.
    rdSMBOMKeyCol_df = rowsBetween[['LVL', 'DPN', 'QPA', 'Agile Description']]
    rdSMBOMKeyCol_df['Agile Description'] = rdSMBOMKeyCol_df['Agile Description'].str.rstrip() #Remove blank at the end of the description.
    rdSMBOMKeyCol_df = rdSMBOMKeyCol_df.reset_index(drop = True)
    
    return rdSMBOMKeyCol_df

###----------Get list of 'Unique Identifier' column when 'LVL' == 0.----------###
def getFullConfig(filePath):
    # TODO May check all columns when LVL == 0
    #Read RD SMBOM.
    wb = pylightxl.readxl(fn = filePath, ws = 'BOM')
    data = []
    for row in wb.ws('BOM').rows:
        row_data = []
        for cell in row:
            if isinstance(cell, str):
                row_data.append(cell)
            elif isinstance(cell, int):
                row_data.append(cell)
            elif isinstance(cell, float):
                row_data.append(cell)
        data.append(row_data)
    rdSMBOM_df = pd.DataFrame(data[1:], columns = data[0])
    rdSMBOM_df['Unique Identifier'] = rdSMBOM_df['Unique Identifier'].str.rstrip()
    lvl_uid_df = rdSMBOM_df[['LVL', 'Unique Identifier']]
    lvlEqual0_df = lvl_uid_df[lvl_uid_df['LVL'] == 0]
    uidList = lvlEqual0_df['Unique Identifier'].values.tolist()
    
    return uidList

###----------Create initial Excel file of change list with specific format.----------###
def initChangeList():
    dateTime = 'Date & Time: '+ time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    global outputPath
    outputPath = os.path.join(os.path.dirname(window.rdFileNameText.toPlainText()), 'Output')
    if os.path.exists(outputPath) == False:
        os.mkdir(outputPath)
    global currTime
    currTime = time.strftime("%Y%m%d-%H%M%S")
    #----------Set initial Excel file with specific format.----------
    with pd.ExcelWriter(outputPath + '/' + currTime + '_Agile update.xlsx', engine = 'xlsxwriter', mode = 'w') as writer: 
        data = {'''
        some columns
                '''
                }
        initExcel_df = pd.DataFrame(data)
        initExcel_df.to_excel(writer, sheet_name = 'Sheet1', index = False)
        wb = writer.book
        ws = writer.sheets['Sheet1']
        for idx, col in enumerate(initExcel_df.columns):
            colWidth = len(col)
            ws.set_column(idx, idx, colWidth)
        '''
        cell format setting
        '''
    
    global logList
    logList = []
    user = 'User: ' + os.getlogin() + '\n'
    inputRD = 'Input RD BOM: ' + window.rdFileNameText.toPlainText()
    inputAgile = 'Input Agile BOM: ' + window.agileFileNameText.toPlainText()
    inputConfig = 'Input Config: ' + window.configInputText.toPlainText() + '\n'
    initFileLog = 'Tool Progress: Running....  5 %,  Total Execution Time is {} seconds'.format(Decimal(str(time.time() - startTime)).quantize(Decimal('0.01'), rounding = ROUND_HALF_UP)) # Round the time string.
    logList.append(dateTime)
    logList.append(user)
    logList.extend([inputRD, inputAgile, inputConfig])
    logList.append('Process:')
    logList.append(initFileLog)
    window.progressBar.setValue(5)
    
###----------Add comparison into change list Excel file.----------###
def toChangeList(agileBOM, rdBOM, config):
    #Remove rows that 'LVL' == 0 in RD SMBOM.
    rdBOM.drop(rdBOM[rdBOM.LVL == 0].index, inplace = True)
    #Load agile update format as a dataframe.
    agileUpdateFmt_df = pd.read_excel(outputPath + '/' +currTime + '_Agile update.xlsx', header = 0, index_col = None, engine = 'openpyxl')
    agileUpdateFmt_df = agileUpdateFmt_df.replace('', np.nan)
    
    #----------Build agile BOM tree.----------
    #Create the root node.
    agileRoot = Node(config, qty = '', desc = '')
    #Initialize stack with root node and level 0.
    agileStack = [(agileRoot, 0)]
    #Iterate through BOM rows.
    for idx, row in agileBOM.iterrows():
        #Create a new node for the row.
        agileNode = Node(row['Number'], parent = None, qty = row['BOM.Qty'], desc = row['*Description'])
        #Check the level of the row and find its parent node.
        while agileStack[-1][1] >= row['Level']:
            agileStack.pop()
        agileParentNode = agileStack[-1][0]
        #Add the new node to the parent node.
        agileNode.parent = agileParentNode
        #Push the new node and its level onto the stack.
        agileStack.append((agileNode, row['Level']))
    #Copy agile BOM tree and add level attribute in nodes.
    agileTreeCopy = copy.deepcopy(agileRoot)
    for node in PreOrderIter(agileTreeCopy):
        node.level = node.depth
    
    #----------Build RD SMBOM tree.----------
    rdRoot = Node(config, qty = '', desc = '')
    rdStack = [(rdRoot, 0)]
    for idx, row in rdBOM.iterrows():
        rdNode = Node(row['DPN'], parent = None, qty = row['QPA'], desc = row['Agile Description'])
        while rdStack[-1][1] >= row['LVL']:
            rdStack.pop()
        rdParentNode = rdStack[-1][0]
        rdNode.parent = rdParentNode
        rdStack.append((rdNode, row['LVL']))
    #Copy RD SMBOM tree and add level attribute in nodes.
    rdTreeCopy = copy.deepcopy(rdRoot)
    for node in PreOrderIter(rdTreeCopy):
        node.level = node.depth
    
    buildTreeLog = 'Tool Progress: Running....  25 %,  Total Execution Time is {} seconds'.format(Decimal(str(time.time() - startTime)).quantize(Decimal('0.01'), rounding = ROUND_HALF_UP))
    logList.append(buildTreeLog)
    window.progressBar.setValue(25)
    
    #----------Add node information into list and convert to dataframe.----------
    agileTreeStruct = []
    agileLevel = []
    agilePN = []
    agileQty = []
    agileDesc = []
    for node in PreOrderIter(agileTreeCopy):
        agileTreeStruct.append(str(node).split(',', 1)[0][7:-1])
        agileLevel.append(int(node.level))
        agilePN.append(str(node.name))
        agileQty.append(node.qty)
        agileDesc.append(str(node.desc))
    agileTree_df = pd.DataFrame(data = agileTreeStruct[0:], columns = ['Tree Structure'])
    agileTree_df = agileTree_df.join(pd.DataFrame(data = agileLevel[0:], columns = ['Level']))
    agileTree_df = agileTree_df.join(pd.DataFrame(data = agilePN[0:], columns = ['PN']))
    agileTree_df = agileTree_df.join(pd.DataFrame(data = agileQty[0:], columns = ['Qty']))
    agileTree_df = agileTree_df.join(pd.DataFrame(data = agileDesc[0:], columns = ['Desc']))
    
    rdTreeStruct = []
    rdLevel = []
    rdPN = []
    rdQty = []
    rdDesc = []
    for node in PreOrderIter(rdTreeCopy):
        rdTreeStruct.append(str(node).split(',', 1)[0][7:-1])
        rdLevel.append(int(node.level))
        rdPN.append(str(node.name))
        rdQty.append(node.qty)
        rdDesc.append(str(node.desc))
    rdTree_df = pd.DataFrame(data = rdTreeStruct[0:], columns = ['Tree Structure'])
    rdTree_df = rdTree_df.join(pd.DataFrame(data = rdLevel[0:], columns = ['Level']))
    rdTree_df = rdTree_df.join(pd.DataFrame(data = rdPN[0:], columns = ['PN']))
    rdTree_df = rdTree_df.join(pd.DataFrame(data = rdQty[0:], columns = ['Qty']))
    rdTree_df = rdTree_df.join(pd.DataFrame(data = rdDesc[0:], columns = ['Desc']))
    
    treeToDFLog = 'Tool Progress: Running....  50 %,  Total Execution Time is {} seconds'.format(Decimal(str(time.time() - startTime)).quantize(Decimal('0.01'), rounding = ROUND_HALF_UP))
    logList.append(treeToDFLog)
    window.progressBar.setValue(50)
    
    #----------Merge the 2 dataframe to compare with each other and add it into agileUpdate_df.----------
    treeMerged_df = pd.merge(rdTree_df, agileTree_df, on = ['Tree Structure', 'Level', 'PN'], how = 'outer', suffixes = ('_new', '_old'), indicator = True)
    treeMerged_df['note'] = ''
    treeMerged_df.fillna('', inplace = True) #Convert NaN to empty string here so that it can use == to check the status.
    #Sort the dataframe to show hierarchy after merging.
    treeMerged_df = treeMerged_df.sort_values(by = ['Tree Structure', 'Level'], kind = 'mergesort', ignore_index = True)
    #Identify the status of rows in 2 dataframes and add data into agile format at the same time.
    for idx, row in treeMerged_df.iterrows():
        if row['_merge'] == 'left_only':
            #Row is present in only new dataframe. -> ADD
            treeMerged_df.loc[idx, 'note'] = 'ADD'
            if int(row['Level']) == 0:
                level = 'Level{}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'ADD'
            elif int(row['Level']) == 1:
                level = 'Level{}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'ADD'
            else:
                level = 'Level {}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'ADD'
        elif row['_merge'] == 'right_only':
            #Row is present in only old dataframe. -> remove
            treeMerged_df.loc[idx, 'note'] = 'remove'
            if int(row['Level']) == 0:
                level = 'Level{}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_old']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_old']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'remove'
            elif int(row['Level']) == 1:
                level = 'Level{}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_old']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_old']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'remove'
            else:
                level = 'Level {}'.format(int(row['Level']))
                agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_old']
                agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_old']
                agileUpdateFmt_df.at[idx, 'BOM note'] = 'remove'
        else:
            #Row is present in both dataframes
            if not row['Qty_new'] == row['Qty_old'] or not row['Desc_new'] == row['Desc_old']:
                #Row is present in both dataframes but with different values. -> Qty for Q'ty ## to ##; Agile Description for update SPEC and writing new above, old below.
                treeMerged_df.loc[idx, 'note'] = 'modified'
                if not row['Qty_new'] == row['Qty_old']: #Qty diff.
                    if int(row['Level']) == 0:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new'])
                    elif int(row['Level']) == 1:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new'])
                    else:
                        level = 'Level {}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new'])
                elif not row['Desc_new'] == row['Desc_old']: #Desc diff.
                    if int(row['Level']) == 0:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'update SPEC'
                    elif int(row['Level']) == 1:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'update SPEC'
                    else:
                        level = 'Level {}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'update SPEC'
                elif not row['Qty_new'] == row['Qty_old'] and not row['Desc_new'] == row['Desc_old']: #Qty and desc diff.
                    if int(row['Level']) == 0:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new']) + ';' + 'update SPEC'
                    elif int(row['Level']) == 1:
                        level = 'Level{}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new']) + ';' + 'update SPEC'
                    else:
                        level = 'Level {}'.format(int(row['Level']))
                        agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                        agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                        agileUpdateFmt_df.at[idx, 'Agile Description'] = 'new:' + row['Desc_new'] + '\n' + 'old:' + row['Desc_old']
                        agileUpdateFmt_df.at[idx, 'BOM note'] = 'Q{}ty {} to {}'.format("'", row['Qty_old'], row['Qty_new']) + ';' + 'update SPEC'
            else:
                #Row is present in both dataframes. -> ''
                treeMerged_df.loc[idx, 'note'] = 'unchange'
                if int(row['Level']) == 0:
                    level = 'Level{}'.format(int(row['Level']))
                    agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                    agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                    agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                    agileUpdateFmt_df.at[idx, 'BOM note'] = ''
                elif int(row['Level']) == 1:
                    level = 'Level{}'.format(int(row['Level']))
                    agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                    agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                    agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                    agileUpdateFmt_df.at[idx, 'BOM note'] = ''
                else:
                    level = 'Level {}'.format(int(row['Level']))
                    agileUpdateFmt_df.at[idx, level] = str(row['PN'])
                    agileUpdateFmt_df.at[idx, 'Qty'] = row['Qty_new']
                    agileUpdateFmt_df.at[idx, 'Agile Description'] = row['Desc_new']
                    agileUpdateFmt_df.at[idx, 'BOM note'] = '' 
    
    treeMerged_df.drop('_merge', axis = 1, inplace = True)
    pd.set_option('display.max_columns', treeMerged_df.columns.size)
    pd.set_option('display.max_rows', None)
    pd.set_option('expand_frame_repr', False) #It will show dataframe columns completely without truncating.
    
    #print(treeMerged_df)
    mergeToCompareLog = 'Tool Progress: Running....  75 %,  Total Execution Time is {} seconds'.format(Decimal(str(time.time() - startTime)).quantize(Decimal('0.01'), rounding = ROUND_HALF_UP))
    logList.append(mergeToCompareLog)
    window.progressBar.setValue(75)
    
    #----------Append the data of agileUpdateFmt_df to initial Excel file.----------
    with pd.ExcelWriter(outputPath + '/' + currTime + '_Agile update.xlsx', engine = 'openpyxl', mode = 'a', if_sheet_exists = 'overlay') as writer:
        agileUpdateFmt_df.to_excel(writer, sheet_name = 'Sheet1', header = False, index = False, startrow = 1, startcol = 0)
        wb = writer.book
        ws = writer.sheets['Sheet1']
        
        #Set the global format first.
        #Format of level 0 row.
        for cell in ws['2:2']:
            cell.font = openpyxl.styles.Font(name = 'Arial', color = 'ffffff', size = 10, bold = True)
            cell.alignment = openpyxl.styles.Alignment(horizontal = 'left', vertical = 'bottom')
            cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = '000000', end_color = '000000')
            thinA2 = openpyxl.styles.Side(border_style = 'thin')
            cell.border = openpyxl.styles.Border(left = thinA2, right = thinA2, top = thinA2, bottom = thinA2)
        #Format of others.
        for row in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row = row, column = col)
                cell.font = openpyxl.styles.Font(name = 'Arial', size = 10)
                if col == 10: #Format for Qty.
                    cell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
                else:
                    cell.alignment = openpyxl.styles.Alignment(horizontal = 'left', vertical = 'center')
                thin = openpyxl.styles.Side(border_style = 'thin')
                cell.border = openpyxl.styles.Border(left = thin, right = thin, top = thin, bottom = thin)
        #Set the format of status.
        for row in range(3, ws.max_row + 1):
            startCol = None
            rowContainsAdd = False
            rowContainsRemove = False
            rowContainsUpdateD = False
            rowContainsUpdateQ = False
            rowContainsUpdateDQ = False
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row = row, column = col)
                if cell.value is not None:
                    if startCol is None:
                        startCol = col
                    if cell.value == 'ADD':
                        rowContainsAdd = True
                        break
                    elif cell.value == 'remove':
                        rowContainsRemove = True
                        break
                    elif cell.value == 'update SPEC':
                        rowContainsUpdateD = True
                        break
                    elif "Q'ty" in str(cell.value):
                        rowContainsUpdateQ = True
                        break
                    elif "Q'ty" in str(cell.value) and 'update SPEC' in str(cell.value):
                        rowContainsUpdateDQ = True
                        break
            if rowContainsAdd: #Change the format of 'ADD' status.
                for col in range(startCol, 19):
                    cell = ws.cell(row = row, column = col)
                    cell.font = openpyxl.styles.Font(name = 'Arial', color = '002060', size = 10, bold = True) #Fonts setting.
                    cell.fill = openpyxl.styles.PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid') #Background color.
            elif rowContainsRemove: #Change the format of 'remove' status.
                for col in range(startCol, 19):
                    cell = ws.cell(row = row, column = col)
                    cell.font = openpyxl.styles.Font(name = 'Arial', strike = True, color = 'FF0000', size = 10)
                    if col == 18: #Strike is no need in the format of 'BOM note'.
                        cell = ws.cell(row = row, column = col)
                        cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
            elif rowContainsUpdateD: #Change the format of 'update SPEC' status.
                for col in range(startCol, 19):
                    cell = ws.cell(row = row, column = col)
                    cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
                    cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = 'FFFF00', end_color = 'FFFF00')
                    if col == 11:
                        cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
                        cell.alignment = openpyxl.styles.Alignment(wrap_text = True) #Separate the update infomation to 2 lines in Agile Description cell.
            elif rowContainsUpdateQ: #Change the format of Q'ty update.
                for col in range(startCol, 19):
                    cell = ws.cell(row = row, column = col)
                    cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
                    cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = 'FFFF00', end_color = 'FFFF00')
                    if col == 10: #Qty needs bold.
                        cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10, bold = True)
            elif rowContainsUpdateDQ: #Change the format of Q'ty update and 'update SPEC'.
                for col in range(startCol, 19):
                    cell = ws.cell(row = row, column = col)
                    cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
                    cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = 'FFFF00', end_color = 'FFFF00')
                    if col == 10:
                        cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10, bold = True)
                    elif col == 11:
                        cell.font = openpyxl.styles.Font(name = 'Arial', color = 'FF0000', size = 10)
                        cell.alignment = openpyxl.styles.Alignment(wrap_text = True)
    
    finLog = 'Tool Progress: Done.  100 %,  Total Execution Time is {} seconds'.format(Decimal(str(time.time() - startTime)).quantize(Decimal('0.01'), rounding = ROUND_HALF_UP))
    logList.append(finLog)
    window.progressBar.setValue(100)

global backupPath
backupPath = os.path.join(r"Path", '{}'.format(os.getlogin() + ' ' + time.strftime("%Y%m%d-%H%M%S")))
#----------Log file.----------
def logFile():
    if os.path.exists(backupPath) == False:
        os.mkdir(backupPath)
    with open(backupPath + '/UsingRecord.txt', 'w') as f:
        for log in logList:
            f.write(log + '\n')
    
if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindowController()
    window.show()
    sys.exit(app.exec_())
